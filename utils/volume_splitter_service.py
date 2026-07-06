"""Business logic for the volume-splitter routes (WP1.7 extraction).

Route handlers in ``routes/volume_splitter.py`` stay thin: they parse/validate
HTTP input, call the helpers here, and shape the ``success_response`` /
``error_response`` envelopes. The logic below is moved verbatim from the
pre-extraction route; behavior (DB queries, formatting, range sanitization,
Excel assembly) is preserved exactly.

Second classification vocabulary (per REFACTOR_PLAN.md WP1.7)
------------------------------------------------------------
The volume-adequacy *status* labels used by this feature -- ``'low'``,
``'optimal'``, ``'high'``, and ``'excessive'`` -- form a SEPARATE
classification vocabulary from the canonical volume *muscle* taxonomy defined
in ``utils/volume_taxonomy.py`` (``BASIC_MUSCLE_GROUPS`` /
``ADVANCED_MUSCLE_GROUPS`` and their mapping tables). These status labels
classify a muscle's weekly-set count against its per-muscle min/max range
(with a ``sets_per_session > 10`` -> ``'excessive'`` override); the status loop
that emits them still lives in ``routes.volume_splitter.calculate_volume``.
Per WP1.7 this vocabulary is documented, NOT consolidated with the canonical
volume-taxonomy classes in this behavior-preserving WP.
"""
import datetime
from io import BytesIO
from typing import cast

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.database import DatabaseHandler
from utils.logger import get_logger
from utils.volume_taxonomy import BASIC_MUSCLE_GROUPS, ADVANCED_MUSCLE_GROUPS

logger = get_logger()


def get_muscle_list_for_mode(mode: str):
    return BASIC_MUSCLE_GROUPS if (mode or "").lower() != "advanced" else ADVANCED_MUSCLE_GROUPS


def build_default_ranges(muscles):
    return {m: {"min": 12, "max": 20} for m in muscles}


def sanitize_range_value(value, fallback):
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return fallback

    if numeric < 0:
        return fallback

    return numeric


def parse_requested_ranges(raw_ranges, muscles):
    defaults = build_default_ranges(muscles)
    if not isinstance(raw_ranges, dict):
        return defaults

    sanitized = {}
    for muscle in muscles:
        fallback = defaults[muscle]
        entry = raw_ranges.get(muscle, fallback)
        if not isinstance(entry, dict):
            sanitized[muscle] = fallback
            continue

        min_value = sanitize_range_value(entry.get("min"), fallback["min"])
        max_value = sanitize_range_value(entry.get("max"), fallback["max"])

        if max_value < min_value:
            max_value = min_value

        sanitized[muscle] = {"min": min_value, "max": max_value}

    return sanitized


def fetch_volume_history():
    """Load the 100 most recent volume plans grouped by plan id.

    Returns the ``{plan_id: {...}}`` mapping the route wraps in
    ``success_response``. Raises on DB error so the route can log/return 500.
    """
    with DatabaseHandler() as db:
        history = db.fetch_all('''
            SELECT vp.id, vp.training_days, vp.created_at, vp.mode, vp.is_active,
                   mv.muscle_group, mv.weekly_sets, mv.sets_per_session, mv.status
            FROM volume_plans vp
            JOIN muscle_volumes mv ON vp.id = mv.plan_id
            ORDER BY vp.created_at DESC
            LIMIT 100
        ''')

    formatted_history = {}
    for row in history:
        plan_id = row['id']
        if plan_id not in formatted_history:
            formatted_history[plan_id] = {
                'training_days': row['training_days'],
                'created_at': row['created_at'],
                'mode': row.get('mode') or 'basic',
                'is_active': bool(row.get('is_active')),
                'muscles': {}
            }
        formatted_history[plan_id]['muscles'][row['muscle_group']] = {
            'weekly_sets': row['weekly_sets'],
            'sets_per_session': row['sets_per_session'],
            'status': row['status']
        }

    return formatted_history


def fetch_volume_plan(plan_id):
    """Load a single volume plan by id, or ``None`` if it does not exist."""
    with DatabaseHandler() as db:
        rows = db.fetch_all('''
            SELECT vp.*, mv.muscle_group, mv.weekly_sets, mv.sets_per_session, mv.status
            FROM volume_plans vp
            JOIN muscle_volumes mv ON vp.id = mv.plan_id
            WHERE vp.id = ?
        ''', (plan_id,))

    if not rows:
        return None

    plan = {
        'training_days': rows[0]['training_days'],
        'created_at': rows[0]['created_at'],
        'mode': rows[0].get('mode') or 'basic',
        'is_active': bool(rows[0].get('is_active')),
        'volumes': {}
    }

    for row in rows:
        plan['volumes'][row['muscle_group']] = {
            'weekly_sets': row['weekly_sets'],
            'sets_per_session': row['sets_per_session'],
            'status': row['status']
        }

    return plan


def delete_volume_plan_record(plan_id):
    """Delete a volume plan by id.

    Returns ``True`` when the plan existed and was deleted, ``False`` when no
    such plan exists (the route maps that to a 404). Raises on DB error.
    """
    with DatabaseHandler() as db:
        existing_plan = db.fetch_one('SELECT id FROM volume_plans WHERE id = ?', (plan_id,))
        if not existing_plan:
            return False

        db.execute_query('DELETE FROM volume_plans WHERE id = ?', (plan_id,))

    return True


def build_volume_excel(data):
    """Build the volume-plan Excel workbook from a request payload.

    Returns ``(BytesIO, download_name)``. The ``training_days`` fallback and
    filename timestamp match the original route behavior exactly.
    """
    try:
        training_days = int(data.get('training_days', 3))
    except (TypeError, ValueError):
        training_days = 3
    training_days = max(training_days, 1)
    volumes = data.get('volumes', {})

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    ws.title = "Volume Plan"

    # Add headers
    headers = ['Muscle Group', 'Sets per Week', 'Sets per Session']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add data
    row = 2
    for muscle, weekly_sets in volumes.items():
        sets_per_session = round(weekly_sets / training_days, 1)
        ws.cell(row=row, column=1, value=muscle)
        ws.cell(row=row, column=2, value=weekly_sets)
        ws.cell(row=row, column=3, value=sets_per_session)
        row += 1

    # Style the worksheet
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 15

    # Create the file
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    return excel_file, f'volume_plan_{timestamp}.xlsx'
