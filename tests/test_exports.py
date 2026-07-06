"""
Tests for data export functionality with focus on:
- Memory safety
- Filename sanitization
- Large dataset handling
- Streaming exports
"""

import os
import pytest
import json
from io import BytesIO
from openpyxl import load_workbook
from utils.export_utils import (
    sanitize_filename,
    create_content_disposition_header,
    generate_timestamped_filename,
    estimate_export_size,
    should_use_streaming,
    create_excel_workbook,
    _remove_temp_file_with_retry,
)


class TestFilenameSanitization:
    """Test filename sanitization for security and safety."""
    
    def test_sanitize_basic_filename(self):
        """Test basic filename sanitization."""
        result = sanitize_filename("workout_summary.xlsx")
        assert result == "workout_summary.xlsx"
    
    def test_sanitize_special_characters(self):
        """Test removal of special characters."""
        result = sanitize_filename("workout<script>.xlsx")
        assert result == "workout_script.xlsx"
        assert "<" not in result
        assert ">" not in result
    
    def test_sanitize_path_traversal(self):
        """Test protection against path traversal attacks."""
        result = sanitize_filename("../../etc/passwd.xlsx")
        assert result == "passwd.xlsx"
        assert ".." not in result
        assert "/" not in result
    
    def test_sanitize_windows_path(self):
        """Test handling of Windows paths."""
        result = sanitize_filename("C:\\Windows\\System32\\file.xlsx")
        assert result == "file.xlsx"
        assert "\\" not in result
    
    def test_sanitize_empty_filename(self):
        """Test fallback for empty filename."""
        result = sanitize_filename("")
        assert result == "export.xlsx"
    
    def test_sanitize_long_filename(self):
        """Test truncation of long filenames."""
        long_name = "a" * 300 + ".xlsx"
        result = sanitize_filename(long_name)
        assert len(result) <= 200
        assert result.endswith(".xlsx")
    
    def test_sanitize_spaces_and_underscores(self):
        """Test handling of spaces and underscores."""
        result = sanitize_filename("my    workout___file.xlsx")
        assert result == "my_workout_file.xlsx"
    
    def test_sanitize_no_extension(self):
        """Test adding xlsx extension when missing."""
        result = sanitize_filename("workout_summary")
        assert result == "workout_summary.xlsx"
    
    def test_sanitize_wrong_extension(self):
        """Test replacing wrong extension."""
        result = sanitize_filename("workout_summary.txt")
        assert result == "workout_summary.xlsx"


class TestContentDisposition:
    """Test Content-Disposition header generation."""
    
    def test_content_disposition_attachment(self):
        """Test attachment disposition."""
        result = create_content_disposition_header("workout.xlsx", attachment=True)
        assert "attachment" in result
        assert "workout.xlsx" in result
    
    def test_content_disposition_inline(self):
        """Test inline disposition."""
        result = create_content_disposition_header("workout.xlsx", attachment=False)
        assert "inline" in result
    
    def test_content_disposition_special_chars(self):
        """Test handling of special characters in filename."""
        result = create_content_disposition_header("workout<test>.xlsx")
        assert "<" not in result
        assert ">" not in result


class TestTimestampedFilename:
    """Test timestamped filename generation."""
    
    def test_generate_timestamped_filename(self):
        """Test basic timestamped filename generation."""
        result = generate_timestamped_filename("workout_summary")
        assert "workout_summary_" in result
        assert result.endswith(".xlsx")
        # Should contain timestamp pattern YYYYMMDD_HHMMSS
        assert len(result.split("_")) >= 3
    
    def test_timestamped_filename_custom_extension(self):
        """Test with custom extension - should still use xlsx."""
        result = generate_timestamped_filename("report", "csv")
        # Should still use xlsx as it's enforced for Excel workbooks
        assert result.endswith(".xlsx")
    
    def test_timestamped_filename_sanitized(self):
        """Test that generated filename is sanitized."""
        result = generate_timestamped_filename("my<test>file")
        assert "<" not in result
        assert ">" not in result


class TestExportSizeEstimation:
    """Test export size estimation for streaming decisions."""
    
    def test_estimate_small_export(self):
        """Test estimation for small export."""
        size = estimate_export_size(100, 10)
        assert size > 0
        assert size < 100000  # Less than 100KB
    
    def test_estimate_large_export(self):
        """Test estimation for large export."""
        size = estimate_export_size(10000, 20)
        assert size > 1000000  # More than 1MB
    
    def test_should_use_streaming_small(self):
        """Test that small exports don't use streaming."""
        result = should_use_streaming(100, 10)
        assert result is False
    
    def test_should_use_streaming_large(self):
        """Test that large exports use streaming."""
        result = should_use_streaming(100000, 50)
        assert result is True


class TestTempFileCleanupRetry:
    """Test the retry-based temp-file cleanup (replaces the old unconditional 0.5s sleep)."""

    def test_removes_file_on_first_try(self, tmp_path):
        """Normal case: no lock, file is removed immediately."""
        target = tmp_path / "sample.xlsx"
        target.write_text("data")

        result = _remove_temp_file_with_retry(str(target))

        assert result is True
        assert not target.exists()

    def test_missing_file_treated_as_success(self, tmp_path):
        """A file that's already gone (e.g. removed elsewhere) is not an error."""
        missing = tmp_path / "already_gone.xlsx"

        result = _remove_temp_file_with_retry(str(missing))

        assert result is True

    def test_recovers_from_transient_lock(self, tmp_path, monkeypatch):
        """Simulates a Windows-style transient PermissionError that clears after a couple tries."""
        import utils.export_utils as export_utils

        target = tmp_path / "locked.xlsx"
        target.write_text("data")

        real_remove = os.remove
        calls = {"count": 0}

        def flaky_remove(path):
            calls["count"] += 1
            if calls["count"] < 3:
                raise PermissionError("simulated transient lock")
            real_remove(path)

        monkeypatch.setattr(export_utils.os, "remove", flaky_remove)
        monkeypatch.setattr(export_utils, "time", _NoSleepTime())

        result = _remove_temp_file_with_retry(str(target), attempts=5, initial_delay=0.01)

        assert result is True
        assert calls["count"] == 3
        assert not target.exists()

    def test_exhausts_retries_and_logs_without_raising(self, tmp_path, monkeypatch, caplog):
        """A permanently locked file logs a warning and returns False instead of raising."""
        import logging
        import utils.export_utils as export_utils

        target = tmp_path / "permanently_locked.xlsx"
        target.write_text("data")

        def always_locked(path):
            raise PermissionError("permanently locked")

        monkeypatch.setattr(export_utils.os, "remove", always_locked)
        monkeypatch.setattr(export_utils, "time", _NoSleepTime())
        caplog.set_level(logging.WARNING, logger="hypertrophy_toolbox")

        result = _remove_temp_file_with_retry(str(target), attempts=3, initial_delay=0.01)

        assert result is False
        assert any(
            "Failed to remove temporary file" in record.getMessage()
            for record in caplog.records
        )


class _NoSleepTime:
    """Stand-in for the `time` module that skips real delays in retry-backoff tests."""

    def sleep(self, seconds):
        pass


class TestWorkbookCleanupPaths:
    """Verify create_excel_workbook cleans up its temp file on both success and error paths."""

    def _spy_on_temp_file(self, monkeypatch):
        """Capture the temp file path xlsxwriter writes to, without changing behavior."""
        import tempfile as tempfile_module
        import utils.export_utils as export_utils

        captured = {}
        original_ntf = tempfile_module.NamedTemporaryFile

        def spy_ntf(*args, **kwargs):
            f = original_ntf(*args, **kwargs)
            captured["path"] = f.name
            return f

        monkeypatch.setattr(export_utils.tempfile, "NamedTemporaryFile", spy_ntf)
        return captured

    def test_success_path_produces_valid_workbook_and_removes_temp_file(self, app, monkeypatch):
        captured = self._spy_on_temp_file(monkeypatch)

        with app.app_context():
            response = create_excel_workbook(
                {"Sheet1": [{"col_a": 1, "col_b": "x"}]}, "test_export.xlsx"
            )

        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.data))
        assert "Sheet1" in wb.sheetnames

        assert "path" in captured
        assert not os.path.exists(captured["path"])

    def test_error_path_still_removes_temp_file(self, app, monkeypatch):
        """If workbook construction fails partway through, the temp file must not leak."""
        import utils.export_utils as export_utils

        captured = self._spy_on_temp_file(monkeypatch)

        def boom(*args, **kwargs):
            raise ValueError("forced failure for test")

        monkeypatch.setattr(export_utils, "_setup_formats", boom)

        with app.app_context():
            with pytest.raises(ValueError):
                create_excel_workbook({"Sheet1": [{"col_a": 1}]}, "test_export.xlsx")

        assert "path" in captured
        assert not os.path.exists(captured["path"])


class TestExportsEndpoints:
    """Test export endpoints with actual requests."""
    
    def test_export_to_excel_structure(self, client, sample_workout_plan):
        """Test full Excel export structure."""
        response = client.get('/export_to_excel')
        
        assert response.status_code == 200
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Verify Content-Disposition header
        assert 'Content-Disposition' in response.headers
        disposition = response.headers['Content-Disposition']
        assert 'attachment' in disposition
        assert 'workout_tracker_summary' in disposition
        assert '.xlsx' in disposition
        
        # Verify the file is valid Excel
        wb = load_workbook(BytesIO(response.data))
        assert wb is not None
        
        # Check for expected sheets
        sheet_names = wb.sheetnames
        assert 'Workout Plan' in sheet_names or len(sheet_names) >= 1
    
    def test_export_to_excel_filename_timestamp(self, client):
        """Test that exported filename includes timestamp."""
        response = client.get('/export_to_excel')
        
        disposition = response.headers['Content-Disposition']
        # Should contain timestamp pattern
        assert '_20' in disposition  # Year prefix
    
    def test_export_summary_with_method(self, client, sample_workout_log):
        """Test summary export with specific method."""
        response = client.post('/export_summary',
                               json={'method': 'Total'},
                               content_type='application/json')
        
        assert response.status_code == 200
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Verify filename includes method
        disposition = response.headers['Content-Disposition']
        assert 'workout_summary_Total' in disposition
    
    def test_export_summary_invalid_method(self, client):
        """Test summary export with invalid method gracefully."""
        response = client.post('/export_summary',
                               json={'method': 'InvalidMethod'},
                               content_type='application/json')
        
        # Should still succeed or return graceful error
        assert response.status_code in [200, 400]
    
    def test_export_to_workout_log_success(self, client, sample_workout_plan):
        """Test exporting workout plan to workout log."""
        response = client.post('/export_to_workout_log')
        
        assert response.status_code == 200
        data = json.loads(response.data)
        assert data['status'] == 'success'
        assert 'message' in data

    def test_export_to_workout_log_rejects_get(self, client, sample_workout_plan):
        """Importing a plan is a write operation and must remain POST-only."""
        response = client.get('/export_to_workout_log')

        assert response.status_code == 405

    def test_export_to_workout_log_preserves_exercise_order(
        self, client, sample_workout_plan
    ):
        """Importing a plan must not repair or otherwise rewrite plan ordering."""
        from utils.database import DatabaseHandler

        with DatabaseHandler() as db:
            db.execute_query("UPDATE user_selection SET exercise_order = NULL")
            before_rows = db.fetch_all(
                "SELECT id, exercise_order FROM user_selection ORDER BY id"
            )
            before_snapshot = [
                (row["id"], row["exercise_order"]) for row in before_rows
            ]

        response = client.post('/export_to_workout_log')

        assert response.status_code == 200
        with DatabaseHandler() as db:
            after_rows = db.fetch_all(
                "SELECT id, exercise_order FROM user_selection ORDER BY id"
            )
        assert [
            (row["id"], row["exercise_order"]) for row in after_rows
        ] == before_snapshot

    @pytest.mark.parametrize(
        ("weight", "rir", "minimum", "maximum"),
        [(0, 0, 8, 8), (1000, 10, 8, 12)],
    )
    def test_export_to_workout_log_accepts_canonical_boundaries(
        self, client, sample_workout_plan, weight, rir, minimum, maximum
    ):
        from utils.database import DatabaseHandler

        with DatabaseHandler() as db:
            db.execute_query(
                "UPDATE user_selection SET weight = ?, rir = ?, "
                "min_rep_range = ?, max_rep_range = ?",
                (weight, rir, minimum, maximum),
            )

        response = client.post('/export_to_workout_log')
        assert response.status_code == 200

    @pytest.mark.parametrize(
        ("column", "value"),
        [
            ("weight", -0.01), ("weight", 1000.01),
            ("rir", -0.01), ("rir", 10.01),
            ("min_rep_range", 13), ("max_rep_range", 7),
        ],
    )
    def test_export_to_workout_log_rejects_invalid_legacy_plan_atomically(
        self, client, sample_workout_plan, column, value
    ):
        from utils.database import DatabaseHandler

        with DatabaseHandler() as db:
            db.execute_query(f"UPDATE user_selection SET {column} = ?", (value,))

        response = client.post('/export_to_workout_log')
        assert response.status_code == 400
        assert response.get_json()["error"]["code"] == "VALIDATION_ERROR"
        with DatabaseHandler() as db:
            row = db.fetch_one("SELECT COUNT(*) AS count FROM workout_log")
        assert row is not None
        assert row["count"] == 0
    
    def test_export_to_workout_log_empty_plan(self, client, clean_database):
        """Test exporting empty workout plan."""
        response = client.post('/export_to_workout_log')
        
        assert response.status_code in [400, 404, 500]  # Accept various error codes
        try:
            data = json.loads(response.data)
            assert data['status'] == 'error'
        except:
            # If JSON parsing fails, that's also acceptable for error responses
            pass
    
    def test_export_large_dataset_all(self, client, large_workout_log):
        """Test streaming export with all data."""
        response = client.post('/export_large_dataset',
                               json={'type': 'all'},
                               content_type='application/json')
        
        assert response.status_code == 200
        assert response.headers['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        # Verify it's a valid Excel file
        wb = load_workbook(BytesIO(response.data))
        assert wb is not None
    
    def test_export_large_dataset_workout_log_only(self, client, large_workout_log):
        """Test streaming export with workout log only."""
        response = client.post('/export_large_dataset',
                               json={'type': 'workout_log'},
                               content_type='application/json')
        
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.data))
        assert 'Workout Log' in wb.sheetnames
    
    def test_export_large_dataset_session_summary_only(self, client, large_workout_log):
        """Test streaming export with session summary only."""
        response = client.post('/export_large_dataset',
                               json={'type': 'session_summary'},
                               content_type='application/json')
        
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.data))
        assert 'Session Summary' in wb.sheetnames


class TestExportMemorySafety:
    """Test that exports handle large datasets without memory issues."""
    
    def test_export_with_many_rows(self, client, database_with_rows):
        """Test export with large number of rows (memory safety)."""
        # This test would populate database with many rows
        # and verify export completes without memory errors
        response = client.get('/export_to_excel')
        
        assert response.status_code == 200
        # If we get here without timeout/memory error, test passes
        assert len(response.data) > 0
    
    def test_streaming_export_memory_usage(self, client, large_workout_log):
        """Test that streaming export doesn't load all data at once."""
        import tracemalloc
        
        tracemalloc.start()
        initial_memory = tracemalloc.get_traced_memory()[0]
        
        response = client.post('/export_large_dataset',
                               json={'type': 'all'},
                               content_type='application/json')
        
        peak_memory = tracemalloc.get_traced_memory()[1]
        tracemalloc.stop()
        
        # Memory usage should be reasonable (less than 100MB increase)
        memory_increase = (peak_memory - initial_memory) / (1024 * 1024)  # MB
        assert response.status_code == 200
        assert memory_increase < 100  # Less than 100MB increase
    
    def test_export_respects_max_rows_limit(self, client, database_with_many_rows):
        """Test that exports respect MAX_EXPORT_ROWS limit."""
        # database_with_many_rows fixture returns the test limit used
        test_max_rows = database_with_many_rows
        
        response = client.get('/export_to_excel')
        
        assert response.status_code == 200
        wb = load_workbook(BytesIO(response.data))
        
        # Check that no sheet exceeds the test MAX_EXPORT_ROWS
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            row_count = sheet.max_row
            assert row_count <= test_max_rows + 1  # +1 for header


class TestExportErrorHandling:
    """Test error handling in export operations."""
    
    def test_export_handles_database_error(self, client, broken_database):
        """Test that export handles database errors gracefully."""
        response = client.get('/export_to_excel')
        
        # Should return an error status
        assert response.status_code in [500, 404]
        # Try to parse as JSON, but it might be HTML error page
        try:
            data = json.loads(response.data)
            assert data['status'] == 'error'
        except:
            # HTML error page is also acceptable
            pass
    
    def test_export_handles_missing_data(self, client, clean_database):
        """Test export with no data returns valid empty Excel."""
        response = client.get('/export_to_excel')
        
        # Should either succeed with empty sheets or return appropriate error
        assert response.status_code in [200, 400]
    
    def test_export_summary_handles_bad_json(self, client):
        """Test export summary with invalid JSON."""
        response = client.post('/export_summary',
                               data='invalid json',
                               content_type='application/json')
        
        # Should return error status
        assert response.status_code in [400, 500]
        # Response might be JSON error or HTML page
        # Both are acceptable for this test

    def test_export_summary_returns_empty_workbook(self, client):
        """Ensure summary export returns a valid workbook even with no data."""
        response = client.post('/export_summary', json={'method': 'Total'})
        assert response.status_code == 200
        assert response.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert int(response.headers.get('Content-Length', '0')) > 0


class TestExcelExportReadOnlyContract:
    """Regression coverage for OD3's read-only Excel export contract."""

    def test_export_to_excel_preserves_exercise_order(self, client, clean_database):
        """Excel export must preserve duplicate and NULL ordering values exactly."""
        from utils.database import DatabaseHandler

        with DatabaseHandler() as db:
            db.execute_query("INSERT OR IGNORE INTO exercises (exercise_name, primary_muscle_group) VALUES ('Bench Press', 'Chest')")
            db.execute_query("INSERT OR IGNORE INTO exercises (exercise_name, primary_muscle_group) VALUES ('Squat', 'Legs')")
            db.execute_query("INSERT OR IGNORE INTO exercises (exercise_name, primary_muscle_group) VALUES ('Deadlift', 'Back')")
            db.execute_query("INSERT INTO user_selection (routine, exercise, sets, min_rep_range, max_rep_range, weight, exercise_order) VALUES ('A', 'Bench Press', 3, 8, 12, 100, 1)")
            db.execute_query("INSERT INTO user_selection (routine, exercise, sets, min_rep_range, max_rep_range, weight, exercise_order) VALUES ('A', 'Squat', 3, 8, 12, 100, 1)")
            db.execute_query("INSERT INTO user_selection (routine, exercise, sets, min_rep_range, max_rep_range, weight, exercise_order) VALUES ('B', 'Deadlift', 3, 8, 12, 100, NULL)")
            before_rows = db.fetch_all(
                "SELECT id, exercise_order FROM user_selection ORDER BY id"
            )
            before_snapshot = [(row["id"], row["exercise_order"]) for row in before_rows]

        response = client.get('/export_to_excel')
        assert response.status_code == 200

        with DatabaseHandler() as db:
            after_rows = db.fetch_all(
                "SELECT id, exercise_order FROM user_selection ORDER BY id"
            )
            after_snapshot = [(row["id"], row["exercise_order"]) for row in after_rows]

        assert after_snapshot == before_snapshot


# Fixtures for tests

@pytest.fixture
def sample_workout_plan(client):
    """Create a sample workout plan for testing."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        db.execute_query(
            """
            INSERT OR IGNORE INTO exercises (
                exercise_name,
                primary_muscle_group,
                secondary_muscle_group,
                tertiary_muscle_group,
                force,
                equipment,
                mechanic,
                utility,
                difficulty
            )
            VALUES ('Bench Press', 'Chest', 'Triceps', 'Shoulders', 'Push', 'Barbell', 'Compound', 'Basic', 'Intermediate')
            """
        )
        db.execute_query("""
            INSERT INTO user_selection (routine, exercise, sets, min_rep_range, max_rep_range, weight)
            VALUES ('A', 'Bench Press', 3, 8, 12, 100)
        """)
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM user_selection")
        db.execute_query("DELETE FROM exercises WHERE exercise_name = 'Bench Press'")


@pytest.fixture
def sample_workout_log(client):
    """Create sample workout log entries."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        for i in range(10):
            db.execute_query("""
                INSERT INTO workout_log (routine, exercise, planned_sets, scored_weight, scored_max_reps)
                VALUES ('A', 'Bench Press', 3, 100, 10)
            """)
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM workout_log")


@pytest.fixture
def large_workout_log(client):
    """Create a large workout log for streaming tests."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        # Insert 1000 records for testing
        for i in range(1000):
            db.execute_query("""
                INSERT INTO workout_log (routine, exercise, planned_sets, scored_weight, scored_max_reps)
                VALUES ('A', 'Bench Press', 3, ?, 10)
            """, (100 + i % 50,))
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM workout_log")


@pytest.fixture
def database_with_rows(client):
    """Create database with many rows for memory testing."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        # Insert 5000 records
        for i in range(5000):
            db.execute_query("""
                INSERT INTO workout_log (routine, exercise, planned_sets, scored_weight, scored_max_reps)
                VALUES ('A', 'Exercise ' || ?, 3, ?, 10)
            """, (i, 100 + i % 100))
    
    yield
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM workout_log")


@pytest.fixture
def database_with_many_rows(client, monkeypatch):
    """Create database with rows exceeding MAX_EXPORT_ROWS."""
    from utils.database import DatabaseHandler
    import utils.export_utils
    
    # Use a small limit for testing (100 rows)
    TEST_MAX_ROWS = 100
    monkeypatch.setattr(utils.export_utils, 'MAX_EXPORT_ROWS', TEST_MAX_ROWS)
    
    # Create more rows than the test limit using batch insert
    rows_to_create = TEST_MAX_ROWS + 50  # 150 rows total
    
    with DatabaseHandler() as db:
        # Use batch insert for speed
        values = ",".join(["('A', 'Exercise', 3, 100, 10)"] * rows_to_create)
        db.execute_query(f"""
            INSERT INTO workout_log (routine, exercise, planned_sets, scored_weight, scored_max_reps)
            VALUES {values}
        """)
    
    yield TEST_MAX_ROWS
    
    # Cleanup
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM workout_log")


@pytest.fixture
def clean_database(client):
    """Ensure database is clean for testing."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        db.execute_query("DELETE FROM user_selection")
        db.execute_query("DELETE FROM workout_log")
    
    yield


@pytest.fixture
def broken_database(client, monkeypatch):
    """Simulate a broken database connection."""
    from utils.database import DatabaseHandler
    
    original_fetch_all = DatabaseHandler.fetch_all
    
    def broken_fetch_all(self, *args, **kwargs):
        raise Exception("Database connection failed")
    
    monkeypatch.setattr(DatabaseHandler, 'fetch_all', broken_fetch_all)
    
    yield
    
    # Restore
    monkeypatch.setattr(DatabaseHandler, 'fetch_all', original_fetch_all)
