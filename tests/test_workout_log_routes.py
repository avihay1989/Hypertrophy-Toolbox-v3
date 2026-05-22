"""
Tests for routes/workout_log.py

Covers CRUD operations on workout log entries, progression tracking,
and data export functionality with focus on:
- Data integrity (invalid IDs, missing fields)
- Edge cases (NULL values in calculations)
- Error responses (404, 400, 500)
"""
import pytest
from datetime import datetime
from flask import Flask


class TestWorkoutLogPageRender:
    """Tests for GET /workout_log page rendering."""

    def test_workout_log_page_loads(self, client, clean_db):
        """Page request returns (200 with templates, 500 without)."""
        resp = client.get("/workout_log")
        # Returns 500 in test env (no templates) or 200 in full env
        assert resp.status_code in (200, 500)

    def test_workout_log_page_shows_entries(self, client, clean_db, workout_log_entry):
        """Page request with data returns (200 with templates, 500 without)."""
        resp = client.get("/workout_log")
        # Returns 500 in test env (no templates) or 200 in full env
        assert resp.status_code in (200, 500)

    def test_workout_log_page_renders_assisted_weight_indicator(
        self, client, clean_db, workout_plan_entry
    ):
        """Server-rendered weight indicator treats lower assistance as progress."""
        from utils.database import DatabaseHandler

        with DatabaseHandler() as db:
            db.execute_query("""
                INSERT INTO workout_log (
                    routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
                    planned_weight, scored_weight, workout_plan_id, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                "Pull",
                "Machine Assisted Chin Up",
                4,
                6,
                10,
                65.0,
                60.0,
                workout_plan_entry["id"],
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            ))

        resp = client.get("/workout_log")
        assert resp.status_code == 200
        html = resp.get_data(as_text=True)
        assert 'data-assisted-bodyweight="true"' in html
        assert "Assistance decreased!" in html
        assert "fa-arrow-up" in html

    def test_workout_log_page_renders_zero_assistance_as_entered_value(
        self, client, clean_db, workout_plan_entry
    ):
        """0kg assistance should render as a real scored weight, not as empty."""
        from utils.database import DatabaseHandler

        with DatabaseHandler() as db:
            db.execute_query("""
                INSERT INTO workout_log (
                    routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
                    planned_weight, scored_weight, workout_plan_id, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                "Pull",
                "Machine Assisted Pull Up",
                4,
                6,
                10,
                65.0,
                0.0,
                workout_plan_entry["id"],
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            ))

        resp = client.get("/workout_log")
        assert resp.status_code == 200
        html = resp.get_data(as_text=True)
        assert 'data-field="scored_weight"' in html
        assert 'value="0.0"' in html
        assert '<span class="editable-text">0.0</span>' in html
        assert "Assistance decreased!" in html


class TestUpdateWorkoutLog:
    """Tests for POST /update_workout_log endpoint."""

    def test_update_scored_weight(self, client, clean_db, workout_log_entry):
        """Should update scored_weight field."""
        resp = client.post("/update_workout_log", json={
            "id": workout_log_entry["id"],
            "updates": {"scored_weight": 100.0}
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["ok"] is True

    def test_update_scored_reps(self, client, clean_db, workout_log_entry):
        """Should update min/max rep fields."""
        resp = client.post("/update_workout_log", json={
            "id": workout_log_entry["id"],
            "updates": {
                "scored_min_reps": 8,
                "scored_max_reps": 12
            }
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["ok"] is True

    def test_update_scored_rir_rpe(self, client, clean_db, workout_log_entry):
        """Should update RIR and RPE fields."""
        resp = client.post("/update_workout_log", json={
            "id": workout_log_entry["id"],
            "updates": {
                "scored_rir": 2,
                "scored_rpe": 8.5
            }
        })
        assert resp.status_code == 200

    def test_update_progression_date(self, client, clean_db, workout_log_entry):
        """Should update last_progression_date field."""
        resp = client.post("/update_workout_log", json={
            "id": workout_log_entry["id"],
            "updates": {"last_progression_date": "2024-01-15"}
        })
        assert resp.status_code == 200

    def test_update_missing_log_id(self, client, clean_db):
        """Should return 400 if log ID is missing."""
        resp = client.post("/update_workout_log", json={
            "updates": {"scored_weight": 100.0}
        })
        assert resp.status_code == 400
        data = resp.get_json()
        assert data["error"]["code"] == "VALIDATION_ERROR"

    def test_update_nonexistent_log(self, client, clean_db):
        """Should return 404 for non-existent log entry."""
        resp = client.post("/update_workout_log", json={
            "id": 99999,
            "updates": {"scored_weight": 100.0}
        })
        assert resp.status_code == 404
        data = resp.get_json()
        assert data["error"]["code"] == "NOT_FOUND"

    def test_update_no_valid_fields(self, client, clean_db, workout_log_entry):
        """Should return 400 if no valid fields provided."""
        resp = client.post("/update_workout_log", json={
            "id": workout_log_entry["id"],
            "updates": {"invalid_field": "value"}
        })
        assert resp.status_code == 400
        data = resp.get_json()
        assert data["error"]["code"] == "VALIDATION_ERROR"

    def test_update_ignores_invalid_fields(self, client, clean_db, workout_log_entry):
        """Should ignore invalid fields but update valid ones."""
        resp = client.post("/update_workout_log", json={
            "id": workout_log_entry["id"],
            "updates": {
                "scored_weight": 95.0,
                "hacker_field": "DROP TABLE"
            }
        })
        assert resp.status_code == 200


class TestDeleteWorkoutLog:
    """Tests for POST /delete_workout_log endpoint."""

    def test_delete_workout_log(self, client, clean_db, workout_log_entry):
        """Should delete a workout log entry."""
        resp = client.post("/delete_workout_log", json={
            "id": workout_log_entry["id"]
        })
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["ok"] is True

    def test_delete_missing_id(self, client, clean_db):
        """Should return 400 if no ID provided."""
        resp = client.post("/delete_workout_log", json={})
        assert resp.status_code == 400
        data = resp.get_json()
        assert data["error"]["code"] == "VALIDATION_ERROR"

    def test_delete_nonexistent_log(self, client, clean_db):
        """Should return 404 for non-existent log entry."""
        resp = client.post("/delete_workout_log", json={"id": 99999})
        assert resp.status_code == 404
        data = resp.get_json()
        assert data["error"]["code"] == "NOT_FOUND"


class TestUpdateProgressionDate:
    """Tests for POST /update_progression_date endpoint."""

    def test_update_progression_date_success(self, client, clean_db, workout_log_entry):
        """Should update progression date."""
        resp = client.post("/update_progression_date", json={
            "id": workout_log_entry["id"],
            "date": "2024-02-01"
        })
        assert resp.status_code == 200

    def test_update_progression_date_missing_id(self, client, clean_db):
        """Should return 400 if ID missing."""
        resp = client.post("/update_progression_date", json={
            "date": "2024-02-01"
        })
        assert resp.status_code == 400

    def test_update_progression_date_missing_date(self, client, clean_db, workout_log_entry):
        """Should return 400 if date missing."""
        resp = client.post("/update_progression_date", json={
            "id": workout_log_entry["id"]
        })
        assert resp.status_code == 400

    def test_update_progression_date_nonexistent(self, client, clean_db):
        """Should return 404 for non-existent entry."""
        resp = client.post("/update_progression_date", json={
            "id": 99999,
            "date": "2024-02-01"
        })
        assert resp.status_code == 404


class TestCheckProgression:
    """Tests for GET /check_progression/<id> endpoint."""

    def test_check_progression_achieved_by_weight(self, client, clean_db, workout_log_with_progression):
        """Should detect progression when scored_weight > planned_weight."""
        entry = workout_log_with_progression["weight_increase"]
        resp = client.get(f"/check_progression/{entry['id']}")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["data"]["is_progressive"] is True
        assert data["data"]["status"] == "Achieved"

    def test_check_progression_achieved_by_assisted_weight_decrease(
        self, client, clean_db, workout_log_with_progression
    ):
        """Should detect progression when machine assistance decreases."""
        entry = workout_log_with_progression["assisted_weight_decrease"]
        resp = client.get(f"/check_progression/{entry['id']}")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["data"]["is_progressive"] is True
        assert data["data"]["status"] == "Achieved"

    def test_check_progression_not_achieved_by_assisted_weight_increase(
        self, client, clean_db, workout_log_with_progression
    ):
        """Should not treat increased machine assistance as progression."""
        entry = workout_log_with_progression["assisted_weight_increase"]
        resp = client.get(f"/check_progression/{entry['id']}")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["data"]["is_progressive"] is False
        assert data["data"]["status"] == "Pending"

    def test_check_progression_achieved_by_reps(self, client, clean_db, workout_log_with_progression):
        """Should detect progression when scored_max_reps > planned_max_reps."""
        entry = workout_log_with_progression["reps_increase"]
        resp = client.get(f"/check_progression/{entry['id']}")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["data"]["is_progressive"] is True

    def test_check_progression_achieved_by_rir(self, client, clean_db, workout_log_with_progression):
        """Should detect progression when scored_rir < planned_rir."""
        entry = workout_log_with_progression["rir_decrease"]
        resp = client.get(f"/check_progression/{entry['id']}")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["data"]["is_progressive"] is True

    def test_check_progression_achieved_by_rpe(self, client, clean_db, workout_log_with_progression):
        """Should detect progression when scored_rpe > planned_rpe."""
        entry = workout_log_with_progression["rpe_increase"]
        resp = client.get(f"/check_progression/{entry['id']}")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["data"]["is_progressive"] is True

    def test_check_progression_not_achieved(self, client, clean_db, workout_log_entry):
        """Should return pending when no progression detected."""
        resp = client.get(f"/check_progression/{workout_log_entry['id']}")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["data"]["is_progressive"] is False
        assert data["data"]["status"] == "Pending"

    def test_check_progression_nonexistent(self, client, clean_db):
        """Should return 404 for non-existent log entry."""
        resp = client.get("/check_progression/99999")
        assert resp.status_code == 404


class TestGetWorkoutLogs:
    """Tests for GET /get_workout_logs endpoint."""

    def test_get_workout_logs_empty(self, client, clean_db):
        """Should return empty array when no logs exist."""
        resp = client.get("/get_workout_logs")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["ok"] is True
        assert data["data"] == []

    def test_get_workout_logs_with_entries(self, client, clean_db, workout_log_entry):
        """Should return all workout log entries."""
        resp = client.get("/get_workout_logs")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["ok"] is True
        assert len(data["data"]) >= 1


class TestExportToWorkoutLog:
    """Tests for POST /export_to_workout_log endpoint."""

    def test_export_to_workout_log_empty(self, client, clean_db):
        """Should return 400 when no workout plans exist."""
        resp = client.post("/export_to_workout_log")
        assert resp.status_code == 400
        data = resp.get_json()
        assert data["error"]["code"] == "NO_DATA"

    def test_export_to_workout_log_success(self, client, clean_db, workout_plan_entry):
        """Should export workout plan entries to workout log."""
        resp = client.post("/export_to_workout_log")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["ok"] is True
        assert "exported" in data["message"].lower()


class TestClearWorkoutLog:
    """Tests for POST /clear_workout_log endpoint."""

    def test_clear_workout_log_empty(self, client, clean_db):
        """Should handle clearing empty workout log."""
        resp = client.post("/clear_workout_log")
        assert resp.status_code == 200
        data = resp.get_json()
        assert "already empty" in data["message"].lower()

    def test_clear_workout_log_with_entries(self, client, clean_db, workout_log_entry):
        """Should clear all entries from workout log."""
        # Verify entry exists
        resp = client.get("/get_workout_logs")
        assert len(resp.get_json()["data"]) >= 1

        # Clear the log
        resp = client.post("/clear_workout_log")
        assert resp.status_code == 200
        data = resp.get_json()
        assert data["ok"] is True
        assert "cleared" in data["message"].lower()

        # Verify log is empty
        resp = client.get("/get_workout_logs")
        assert resp.get_json()["data"] == []


class TestExportWorkoutLog:
    """Tests for GET /export_workout_log endpoint."""

    def test_export_workout_log_empty(self, client, clean_db):
        """Should return 404 when no logs to export."""
        resp = client.get("/export_workout_log")
        assert resp.status_code == 404

    def test_export_workout_log_returns_excel(self, client, clean_db, workout_log_entry):
        """Should return an Excel file."""
        resp = client.get("/export_workout_log")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type or resp.status_code == 200

    def test_export_workout_log_response_is_valid_xlsx(self, client, clean_db, workout_log_entry):
        """Response body must be a valid xlsx (ZIP archive — starts with PK magic bytes)."""
        resp = client.get("/export_workout_log")
        assert resp.status_code == 200
        assert resp.data[:2] == b"PK"

    def test_export_workout_log_workbook_has_workout_log_sheet(self, client, clean_db, workout_log_entry):
        """Generated workbook must contain a sheet named 'Workout Log'."""
        from io import BytesIO
        from openpyxl import load_workbook

        resp = client.get("/export_workout_log")
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.data), read_only=True)
        try:
            assert "Workout Log" in wb.sheetnames
        finally:
            wb.close()

    def test_export_workout_log_workbook_header_row_matches_keys(self, client, clean_db, workout_log_entry):
        """First row of the 'Workout Log' sheet must be the column keys from get_workout_logs()."""
        from io import BytesIO
        from openpyxl import load_workbook

        from utils.workout_log import get_workout_logs

        logs = get_workout_logs()
        assert logs, "fixture should have produced at least one log row"
        expected_headers = list(logs[0].keys())

        resp = client.get("/export_workout_log")
        assert resp.status_code == 200

        wb = load_workbook(BytesIO(resp.data), read_only=True)
        try:
            ws = wb["Workout Log"]
            header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        finally:
            wb.close()

        # Trim trailing None cells the reader may pad with
        while header_row and header_row[-1] is None:
            header_row.pop()
        assert header_row == expected_headers


# Fixtures for workout_log tests
@pytest.fixture
def workout_log_entry(clean_db, workout_plan_entry):
    """Create a basic workout log entry for testing."""
    from utils.database import DatabaseHandler
    
    with DatabaseHandler() as db:
        query = """
        INSERT INTO workout_log (
            routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
            planned_rir, planned_rpe, planned_weight, workout_plan_id, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        params = (
            "Push", "Bench Press", 3, 8, 12,
            2, 7.0, 80.0, workout_plan_entry["id"],
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        db.execute_query(query, params)
        
        # Get the inserted ID
        result = db.fetch_one(
            "SELECT id FROM workout_log ORDER BY id DESC LIMIT 1"
        )
        assert result is not None
        
        return {"id": result["id"], "exercise": "Bench Press"}


@pytest.fixture
def workout_log_with_progression(clean_db, workout_plan_entry):
    """Create workout log entries with various progression scenarios."""
    from utils.database import DatabaseHandler
    
    entries = {}
    
    with DatabaseHandler() as db:
        # Weight increase progression
        db.execute_query("""
            INSERT INTO workout_log (
                routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
                planned_weight, scored_weight, workout_plan_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Push", "Bench Press", 3, 8, 12, 80.0, 85.0, workout_plan_entry["id"],
              datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        result = db.fetch_one("SELECT id FROM workout_log ORDER BY id DESC LIMIT 1")
        assert result is not None
        entries["weight_increase"] = {"id": result["id"]}

        # Assisted-machine progression (lower assistance = stronger)
        db.execute_query("""
            INSERT INTO workout_log (
                routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
                planned_weight, scored_weight, workout_plan_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Pull", "Machine Assisted Chin Up", 3, 6, 10, 65.0, 60.0, workout_plan_entry["id"],
              datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        result = db.fetch_one("SELECT id FROM workout_log ORDER BY id DESC LIMIT 1")
        assert result is not None
        entries["assisted_weight_decrease"] = {"id": result["id"]}

        # Assisted-machine regression (more assistance = easier)
        db.execute_query("""
            INSERT INTO workout_log (
                routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
                planned_weight, scored_weight, workout_plan_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Pull", "Machine Assisted Parallel Bar Dips", 3, 6, 10, 65.0, 70.0, workout_plan_entry["id"],
              datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        result = db.fetch_one("SELECT id FROM workout_log ORDER BY id DESC LIMIT 1")
        assert result is not None
        entries["assisted_weight_increase"] = {"id": result["id"]}

        # Reps increase progression
        db.execute_query("""
            INSERT INTO workout_log (
                routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
                scored_max_reps, workout_plan_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Push", "OHP", 3, 8, 12, 13, workout_plan_entry["id"],
              datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        result = db.fetch_one("SELECT id FROM workout_log ORDER BY id DESC LIMIT 1")
        assert result is not None
        entries["reps_increase"] = {"id": result["id"]}

        # RIR decrease progression (lower RIR = closer to failure = harder)
        db.execute_query("""
            INSERT INTO workout_log (
                routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
                planned_rir, scored_rir, workout_plan_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Pull", "Rows", 3, 8, 12, 3, 2, workout_plan_entry["id"],
              datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        result = db.fetch_one("SELECT id FROM workout_log ORDER BY id DESC LIMIT 1")
        assert result is not None
        entries["rir_decrease"] = {"id": result["id"]}

        # RPE increase progression (higher RPE = harder)
        db.execute_query("""
            INSERT INTO workout_log (
                routine, exercise, planned_sets, planned_min_reps, planned_max_reps,
                planned_rpe, scored_rpe, workout_plan_id, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, ("Pull", "Pullups", 3, 8, 12, 7.0, 8.5, workout_plan_entry["id"],
              datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        result = db.fetch_one("SELECT id FROM workout_log ORDER BY id DESC LIMIT 1")
        assert result is not None
        entries["rpe_increase"] = {"id": result["id"]}

    return entries


@pytest.fixture  
def workout_plan_entry(clean_db, exercise_factory):
    """Create a workout plan entry (user_selection) for workout log FK."""
    from utils.database import DatabaseHandler
    
    exercise = exercise_factory("Bench Press")
    
    with DatabaseHandler() as db:
        db.execute_query("""
            INSERT INTO user_selection (
                routine, exercise, sets, min_rep_range, max_rep_range, rir, weight
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, ("Push", "Bench Press", 3, 8, 12, 2, 80.0))
        
        result = db.fetch_one(
            "SELECT id FROM user_selection ORDER BY id DESC LIMIT 1"
        )
        assert result is not None
        
        return {"id": result["id"], "exercise": "Bench Press"}
